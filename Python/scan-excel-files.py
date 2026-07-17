import os
import re
import sys
from openpyxl import Workbook, load_workbook

# 兼容旧版xls文件
try:
    import xlrd
except ImportError:
    xlrd = None

# ===================== 配置区 =====================
SCAN_ROOT_PATH = r"C:/"          # 要扫描的根目录
OUTPUT_EXCEL = "敏感数据扫描结果.xlsx"  # 输出Excel文件名
SCAN_XLSX = True                 # 是否扫描 .xlsx
SCAN_XLS = False                 # 是否扫描 .xls（需安装xlrd）
# ==================================================

# ---------- 规则定义 ----------
def luhn_check(card_num: str) -> bool:
    """银行卡号Luhn校验"""
    digits = [int(ch) for ch in card_num if ch.isdigit()]
    if len(digits) < 16 or len(digits) > 19:
        return False
    parity = len(digits) % 2
    total = 0
    for i, d in enumerate(digits):
        if i % 2 == parity:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    return total % 10 == 0

# 每个规则：名称，正则表达式，可选校验函数（接收匹配字符串，返回bool）
RULES = [
    ("手机号", re.compile(r"1[3-9]\d{9}"), None),
    ("身份证", re.compile(r"[1-9]\d{5}(19|20)\d{2}(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])\d{3}[\dXx]"), None),
    # 统一社会信用代码：增加 \b 边界，确保为18位完整编码
    ("统一社会信用代码", re.compile(r"\b[0-9A-HJ-NPQRTUWXY]{2}\d{6}[0-9A-HJ-NPQRTUWXY]{10}\b"), None),
    ("邮箱", re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"), None),
    ("银行卡号", re.compile(r"\b\d{16,19}\b"), luhn_check),
    # ("固定电话", re.compile(r"\b0\d{2,3}-\d{7,8}\b"), None),
    # 已删除：企业名称、IP地址
]

def check_text(text: str):
    """对文本进行所有规则匹配，返回 [(类型, 匹配内容), ...]"""
    if not isinstance(text, str):
        text = str(text) if text is not None else ""
    hits = []
    for name, pat, validator in RULES:
        for m in pat.findall(text):
            if validator and not validator(m):
                continue
            hits.append((name, m))
    return hits

# ---------- 扫描不同文件类型 ----------
def scan_xlsx(file_path):
    """扫描 .xlsx 文件"""
    hits = []
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for col_idx, cell_val in enumerate(row, start=1):
                    if cell_val is None:
                        continue
                    found = check_text(str(cell_val))
                    for typ, content in found:
                        hits.append({
                            "file": file_path,
                            "location": f"工作表：{sheet_name} 行：{row_idx} 列：{col_idx}",
                            "type": typ,
                            "content": content
                        })
        wb.close()
    except Exception as e:
        print(f"读取失败 {file_path}，错误：{str(e)}")
    return hits

def scan_xls(file_path):
    """扫描 .xls 文件"""
    if xlrd is None:
        print("未安装xlrd，跳过xls文件")
        return []
    hits = []
    try:
        wb = xlrd.open_workbook(file_path)
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                for col_idx, cell_val in enumerate(row, start=1):
                    if cell_val is None:
                        continue
                    found = check_text(str(cell_val))
                    for typ, content in found:
                        hits.append({
                            "file": file_path,
                            "location": f"工作表：{sheet_name} 行：{row_idx+1} 列：{col_idx}",
                            "type": typ,
                            "content": content
                        })
    except Exception as e:
        print(f"读取失败 {file_path}，错误：{str(e)}")
    return hits

def traverse_folder(root_dir):
    """递归遍历文件夹，扫描指定类型文件"""
    all_hits = []
    for root, dirs, files in os.walk(root_dir):
        for fname in files:
            full_path = os.path.join(root, fname)
            lower_name = fname.lower()
            if SCAN_XLSX and lower_name.endswith(".xlsx"):
                print(f"扫描：{full_path}")
                all_hits.extend(scan_xlsx(full_path))
            elif SCAN_XLS and lower_name.endswith(".xls"):
                print(f"扫描：{full_path}")
                all_hits.extend(scan_xls(full_path))
    return all_hits

def export_to_excel(hits, output_path):
    """将命中结果导出为Excel，两列：文件路径，问题描述"""
    wb = Workbook()
    ws = wb.active
    ws.title = "敏感数据"
    ws.append(["文件路径", "问题描述"])
    for item in hits:
        desc = f"{item['location']} | {item['type']}：{item['content']}"
        ws.append([item['file'], desc])
    wb.save(output_path)
    print(f"\n结果已导出至：{output_path}")

if __name__ == "__main__":
    # 重新配置控制台编码（Windows下避免乱码）
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')

    print("开始扫描，请稍候...")
    result = traverse_folder(SCAN_ROOT_PATH)

    if not result:
        print("\n扫描完成，未发现敏感数据。")
    else:
        print(f"\n扫描完成，共发现 {len(result)} 条敏感数据记录。")
        export_to_excel(result, OUTPUT_EXCEL)